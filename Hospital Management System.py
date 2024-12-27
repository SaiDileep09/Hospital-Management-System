import pandas
import os
import openpyxl
print("DILEEP HOSPITAL")
z = os.path.exists("HMSEXCEL.xlsx")
if z:
 wb = openpyxl.load_workbook("HMSEXCEL.xlsx")
else:
 wb = openpyxl.Workbook()
 wb.save("HMSEXCEL.xlsx")
 sh = wb.active
 sh.title = "doctordetails"
 wb["doctordetails"].append(["Doctorname", "Specilization",
"Experience"])
 wb.create_sheet(title="coworkersdetails")
 wb["coworkersdetails"].append(["Name", "Qualification", "Position"])
 wb.create_sheet(title="paitentdetails")
 wb["paitentdetails"].append(["Paitentname", "Gender", "Age",
"Address"])
 wb.save("HMSEXCEL.xlsx")
sheet1 = wb["doctordetails"]
sheet2 = wb["coworkersdetails"]
sheet3 = wb["paitentdetails"]
def display():
 try:
 xls = pandas.ExcelFile("HMSEXCEL.xlsx")
 data = pandas.read_excel(xls, "doctordetails")
 data1 = pandas.read_excel(xls, "coworkersdetails")
 data2 = pandas.read_excel(xls, "paitentdetails")
 print("\t\t1. Doctors Details\n\t\t2. Co-worker Details\n\t\t3.
Patient Details")
 a = int(input("enter your choice: "))
 if a == 1:
 print(data)
 elif a == 2:
 print(data1)
 elif a == 3:
 print(data2)
 else:
 print("Sorry, Your entered wrong choice")
 except ValueError:
 print("OPPS,You entered letters or symbols,Try again")
def enter():
 try:
 print("\t\t1. Doctors Details\n\t\t2. Co-worker Details\n\t\t3.
Patient Details")
 c = int(input("Enter your Choice:"))
 if c == 1:
 di = input("Enter the Doctor Name\n\t").strip()
 dj = input("Enter the specilization\n\t").strip()
 dk = input("Enter your experince\n\t").strip()
 sheet1.append([di, dj, dk])
 wb.save("HMSEXCEL.xlsx")
 elif c == 2:
 cn = input("Enter your Name\n\t").strip()
 cq = input("Enter Your Qualificationn\n\t").strip()
 cp = input("Enter your position\n\t").strip()
 sheet2.append([cn, cq, cp])
 wb.save("HMSEXCEL.xlsx")
 elif c == 3:
 pn = input("enter the paitent name\n\t").strip()
 pa = input("enter the paitent age\n\t").strip()
 pg = input("enter the Gender\n\t").strip()
 pd = input("enter the address\n\t").strip()
 pp = input("enter the phone no\n\t").strip()
 sheet3.append([pn, pg, pa, pd, pp])
 wb.save("HMSEXCEL.xlsx")
 else:
 print("Sorry, Your entered wrong choice")
 except ValueError:
 print("Your entered letters, Try again....")
e = 1
while e != 0:
 print("""
 1. Display the details
 2. Add a new member
 3. Make an exit
 """)
 try:
 b = int(input("Enter your Choice:"))
 if b == 1:
 display()
 elif b == 2:
 enter()
 elif b == 3:
 e = 0
 else:
 print("You entered wrong choice, Please Try Again")
 except ValueError:
 print("You entered wrong choice, Please Try Again")